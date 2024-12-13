from django.shortcuts import render, redirect, get_object_or_404
from .models import *
from .forms import *
from django.contrib import messages
from django.http import JsonResponse
from apps.history.models import *
from apps.finance.models import *
from django.forms.models import model_to_dict
import json


def list_(request):
    search_query = request.GET.get('search', '')

    if search_query:
        products = Product.objects.filter(name__icontains=search_query)
    else:
        products = Product.objects.all()

    context = {
        'products': products
    }
    return render(request, 'product/list.html', context)


def create(request):
    form = ProductForm()
    if request.method == 'POST':
        form = ProductForm(request.POST, request.FILES)
        if form.is_valid():
            product = form.save(commit=False)
            product.shop = request.user.shop
            form.save()
            messages.success(request, 'Товар успешно создан')
            return redirect('product-create')
        else:
            messages.error(request, 'Ошибка при создании товара')
    return render(request, 'product/create.html', {'form': form})

def product_create(request):
    # Проверяем, что запрос является AJAX и имеет метод POST
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        # Преобразуем данные JSON из запроса
        data = json.loads(request.body)
        
        # Создаем объект Product с полученными данными и значениями по умолчанию
        product = Product(
            name=data.get('name'),
            brand=data.get('brand'),
            category=data.get('category'),
            bar_code=data.get('barcode'),
            price=data.get('price', 0),       # Цена по умолчанию 0
            quantity=data.get('quantity', 0), # Количество по умолчанию 0
            sale_price=data.get('sale_price', 0), # Продажная цена по умолчанию 0
            shop=request.user.shop
        )

        # Сохраняем объект и возвращаем ответ JSON
        try:
            product.save()
            return JsonResponse({'success': True, 'barcode': product.bar_code})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Invalid request'})


def details(request, pk):
    product = Product.objects.get(id=pk)

    context = {
        'product': product,
    }
    return render(request, 'product/details.html', context)


def update(request, pk):
    product = Product.objects.get(id=pk)
    form = ProductForm(instance=product)
    if request.method == 'POST':
        form = ProductForm(request.POST, request.FILES, instance=product)
        if form.is_valid():
            form.save()
            return redirect('product-list')

    context = {
        'product': product,
        'form': form
    }
    return render(request, 'product/update.html', context)


def delete(request, pk):
    Product.objects.get(id=pk).delete()
    return redirect('product-list')


def sale(request):
    return render(request, 'product/sale.html')

def income(request):
    return render(request, 'product/income.html')


def get_product(request):
    try:
        bar_code = request.GET.get('bar_code', '')
        if not bar_code:
            return JsonResponse({'error': 'Штрих-код не указан'}, status=400)

        # Находим продукт по штрих-коду
        product = get_object_or_404(Product, bar_code=bar_code)

        # Формируем сериализуемый ответ
        response_data = {
            'id': product.id,
            'name': product.name,
            'bar_code': product.bar_code,
            'quantity': product.quantity,
            'price': product.price,
            'sale_price': product.sale_price,
            'brand': str(product.brand),  # Преобразуем объект Product_brand в строку
            'image': product.image.url if product.image else None,
        }

        return JsonResponse(response_data)
    except Exception as e:
        print(f"Ошибка в get_product: {e}")
        return JsonResponse({'error': 'Внутренняя ошибка сервера'}, status=500)


def create_sell_history(request):
    if request.method == 'POST':
        products = json.loads(request.POST.get('products'))
        amount = request.POST.get('amount')
        change = request.POST.get('change')
        discount = request.POST.get('discount')

        profit = Decimal(0)

        if discount == '':
            discount = 0

        order = OrderHistory.objects.create(
            amount=amount,
            change=change,
            discount=discount,
            shop=request.user.shop
        )
        for item in products:
            product = Product.objects.get(id=item['id'])
            quantity = int(item['quantity'])

            # Проверяем доступность количества
            if product.quantity < quantity:
                return JsonResponse({
                    'status': 'error',
                    'message': f"Недостаточно товара для продукта {product.name}. Доступно: {product.quantity}, запрошено: {quantity}."
                })

            # Сохраняем текущую цену продукта
            SoldHistory.objects.create(
                order=order,
                product=product,
                quantity=quantity,
                price_at_sale=product.sale_price  # Сохраняем цену на момент продажи
            )

            product_profit = (Decimal(product.sale_price) - Decimal(product.price)) * quantity
            profit += product_profit


            # Обновляем количество продукта
            product.quantity -= quantity
            product.save()
        
        profit = profit - Decimal(discount)

        order.profit = profit
        order.save()

        return JsonResponse({'status': 'success'})

def create_income_history(request):
    if request.method == 'POST':
        products = json.loads(request.POST.get('products'))
        amount = request.POST.get('amount')

        # Создаем запись о поступлении
        income = Income.objects.create(
            amount=amount,
            shop=request.user.shop
        )
        expend = Expend.objects.create(
            expend_type='supplies',
            description='Поступление товара',
            start_date=timezone.now(),
            end_date=timezone.now(),
            amount=amount,
            shop=request.user.shop,
            status='paid',
        )

        for item in products:
            product = Product.objects.get(id=item['id'])
            quantity = int(item['quantity'])
            price = float(item['price'])  # Получаем цену продукта из запроса
            sale_price = float(item['sale_price'])  # Получаем продажную цену из запроса

            
            IncomeHistory.objects.create(
                income=income,
                product=product,
                quantity=quantity,
                price_at_income=price
            )

            # Обновляем количество и цены продукта
            product.quantity += quantity
            product.price = price  # Обновляем цену
            product.sale_price = sale_price  # Обновляем продажную цену
            product.save()

        return JsonResponse({'status': 'success'})


def update_quantity(request):
    if request.method == 'POST':
        product_id = request.POST.get('product_id')
        quantity = request.POST.get('quantity')

        product = Product.objects.get(id=product_id)
        product.quantity = quantity
        product.save()

        return JsonResponse({'status': 'success'})


def find_product(request):
    barcode = request.GET.get('barcode')
    try:
        product = Product.objects.get(bar_code=barcode)
        return JsonResponse({'product_name': product.name, 'price': product.price, 'image': product.image.url})
    except Product.DoesNotExist:
        return JsonResponse({'error': 'Product not found'}, status=404)


def search_product(request):
    query = request.GET.get('query', '')
    products = Product.objects.filter(name__icontains=query).values('bar_code', 'name', 'quantity')
    return JsonResponse(list(products), safe=False)

def reports(request):
    # Отчеты
    zero_stock_products = Product.objects.filter(quantity=0)  # Нулевые остатки
    low_stock_products = Product.objects.filter(quantity__lte=3)  # Остатки 3 или меньше
    products_by_category_and_brand = Product.objects.all().order_by('category', 'brand')  # Группировка

    context = {
        'zero_stock_products': zero_stock_products,
        'low_stock_products': low_stock_products,
        'products_by_category_and_brand': products_by_category_and_brand,
    }
    return render(request, 'product/reports.html', context)

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# Отчет по нулевым остаткам
def export_zero_stock(request):
    products = Product.objects.filter(quantity=0)
    wb = Workbook()
    ws = wb.active
    ws.title = 'Zero Stock'

    # Заголовки
    ws.append(['Товар', 'Штрих код', 'Магазин'])

    # Данные
    for product in products:
        ws.append([product.name, product.bar_code, ])

    # Ответ с заголовками для скачивания
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=zero_stock_report.xlsx'
    wb.save(response)
    return response

# Отчет по товарам с остатками < 3
def export_low_stock(request):
    products = Product.objects.filter(quantity__lt=3)
    wb = Workbook()
    ws = wb.active
    ws.title = 'Low Stock'

    # Заголовки
    ws.append(['Товар', 'Штрих код', 'Количество', 'Магазин'])

    # Данные
    for product in products:
        ws.append([product.name, product.bar_code, product.quantity, ])

    # Ответ с заголовками для скачивания
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=low_stock_report.xlsx'
    wb.save(response)
    return response

# Прайс лист по категориям
def export_by_category(request):
    products = Product.objects.all().order_by('category')
    wb = Workbook()
    ws = wb.active
    ws.title = 'Price by Category'

    # Заголовки
    ws.append(['Категория', 'Товар', 'Цена покупки', 'Цена для продажи', 'Магазин'])

    # Данные
    for product in products:
        category_name = product.category.name if product.category else 'Без категории'  # Проверка на наличие категории
        ws.append([category_name, product.name, product.price, product.sale_price])

    # Ответ с заголовками для скачивания
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=price_by_category.xlsx'
    wb.save(response)
    return response

# Прайс лист по брэндам
def export_by_brand(request):
    products = Product.objects.all().order_by('brand')
    wb = Workbook()
    ws = wb.active
    ws.title = 'Price by Brand'

    # Заголовки
    ws.append(['Бренд', 'Товар', 'Цена покупки', 'Цена для продажи', 'Магазин'])

    # Данные
    for product in products:
        brand_name = product.brand.name if product.brand else 'Без бренда'
        ws.append([brand_name, product.name, product.price, product.sale_price, ])

    # Ответ с заголовками для скачивания
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=price_by_brand.xlsx'
    wb.save(response)
    return response

# Общий прайс лист
def export_all(request):
    products = Product.objects.all()
    wb = Workbook()
    ws = wb.active
    ws.title = 'General Price List'

    # Заголовки
    ws.append(['Категория', 'Бренд', 'Товар', 'Цена покупки', 'Цена для продажи', 'Количество', 'Магазин'])

    # Данные
    for product in products:
        category_name = product.category.name if product.category else 'Без категории'  # Проверка на наличие категории
        brand_name = product.brand.name if product.brand else 'Без бренда'
        ws.append([category_name, brand_name, product.name, product.price, product.sale_price, product.quantity, ])

    # Ответ с заголовками для скачивания
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=general_price_list.xlsx'
    wb.save(response)
    return response


def report_selection(request):
    if request.method == 'POST':
        form = ReportSelectionForm(request.POST)
        if form.is_valid():
            category = form.cleaned_data.get('category')
            brand = form.cleaned_data.get('brand')
            
            # Фильтруем продукты по категории или бренду
            if category:
                products = Product.objects.filter(category=category)
            elif brand:
                products = Product.objects.filter(brand=brand)
            else:
                products = Product.objects.all()
            
            # Генерируем отчет
            wb = Workbook()
            ws = wb.active
            ws.title = 'Отчет'

            # Заголовки
            ws.append(['Товар', 'Категория', 'Бренд', 'Цена покупки', 'Цена для продажи'])

            # Заполняем данными
            for product in products:
                ws.append([product.name, product.category.name if product.category else 'Без категории',
                           product.brand.name if product.brand else 'Без бренда',
                           product.price, product.sale_price])

            # Создаем файл Excel и возвращаем его пользователю
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename=report(отчет)_{datetime.now().strftime("%d%m%Y")}.xlsx'
            wb.save(response)
            return response
    else:
        form = ReportSelectionForm()

    return render(request, 'product/report_selection.html', {'form': form})
