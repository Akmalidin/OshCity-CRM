from django.shortcuts import render, redirect
from django.contrib.admin.models import LogEntry
from .models import *
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
import pandas as pd
import openpyxl

@login_required
def history(request):
    orders = OrderHistory.objects.all().order_by('-created')

    incomes = Income.objects.all().order_by('-created')

    history_items = list(orders) + list(incomes)

    history_items.sort(key=lambda x: x.created, reverse=True)

    context = {
        'history_items': history_items,
    }

    return render(request, 'history/history.html', context)


def sales(request):
    sales = SoldHistory.objects.all().order_by('-created')
    context = {
        'sales': sales
    }
    return render(request, 'history/sales.html', context)


def incomes(request):
    incomes = IncomeHistory.objects.all().order_by('-created')
    context = {
        'incomes': incomes
    }
    return render(request, 'history/incomes.html', context)


def sales_delete(request, pk):
    sale = SoldHistory.objects.get(id=pk)
    sale.product.quantity += sale.quantity
    sale.product.save()
    sale.delete()
    return redirect('sold-history')


def incomes_delete(request, pk):
    income = IncomeHistory.objects.get(id=pk)
    product = income.product
    product.quantity -= income.quantity
    product.save()
    income.delete()
    return redirect('income-history')

def order_delete(request, pk):
    order = OrderHistory.objects.get(id=pk)
    for i in order.soldhistory_set.all():
        product = i.product
        product.quantity += i.quantity
        product.save()
    order.delete()
    return redirect('history')

def income_delete(request, pk):
    income = Income.objects.get(id=pk)
    for i in income.incomehistory_set.all():
        product = i.product
        product.quantity -= i.quantity
        product.save()
    income.delete()
    return redirect('history')

def export_history_to_excel(request):
    wb = openpyxl.Workbook()

    # Лист для истории поступлений
    income_sheet = wb.active
    income_sheet.title = 'История поступлений'

    income_sheet.append(['#', 'Товар', 'Итого', 'Принято', 'Скидка', 'Сдача', 'Дата'])

    # Извлекаем все записи из истории поступлений
    income_history = IncomeHistory.objects.all()
    for item in income_history:
        created_time = localtime(item.created).replace(tzinfo=None)
        product_name = item.product.name if item.product else 'Нет товара'
        income_sheet.append([item.id, product_name, item.total_sum(), item.amount, item.discount, item.change, created_time])

    # Лист для истории продаж
    sales_sheet = wb.create_sheet(title="История продаж")
    sales_sheet.append(['#', 'Итого', 'Принято', 'Скидка', 'Сдача', 'Дата', 'Товары'])

    # Извлекаем все записи из истории продаж
    sales_history = SoldHistory.objects.all()
    for item in sales_history:
        created_time = localtime(item.created).replace(tzinfo=None)
        discount = item.discount if item.discount else 0
        change_value = item.change if hasattr(item, 'change') else 0
        
        # Получение информации о товаре
        product_name = item.product.name if item.product else 'Нет товара'

        sales_sheet.append([item.id, item.total_sum(), item.price_at_sale, discount, change_value, created_time, product_name])

    # Отправляем файл Excel в ответ
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="incomes_and_sales_history.xlsx"'

    # Сохраняем книгу в ответ
    wb.save(response)
    return response


from openpyxl import Workbook
from django.http import HttpResponse
from django.utils.timezone import localtime
def export_sales_history_to_excel(request):
    # Создаем новую книгу Excel
    wb = Workbook()

    # Лист для истории продаж (SoldHistory)
    sales_sheet = wb.active
    sales_sheet.title = 'История продаж'

    # Заголовки для таблицы продаж
    sales_sheet.append(['#', 'Товар', 'Количество', 'Цена', 'Итого', 'Скидка', 'Дата'])

    # Извлекаем все записи истории продаж
    sales_history = SoldHistory.objects.all()

    # Проходим по всем записям истории продаж и добавляем их в лист
    for item in sales_history:
        # Убираем временную зону с даты (если она есть)
        created_time = localtime(item.created).replace(tzinfo=None)  # Убираем tzinfo
        # Добавляем строку с информацией о товаре и данных
        sales_sheet.append([
            item.id,
            item.product.name,  # Имя товара
            item.quantity,
            item.price_at_sale,
            item.total_sum(),  # Итоговая сумма (количество * цена)
            item.discount,
            created_time
        ])

    # Отправляем файл Excel в ответ
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="sales_history.xlsx"'

    # Сохраняем книгу в ответ
    wb.save(response)
    return response